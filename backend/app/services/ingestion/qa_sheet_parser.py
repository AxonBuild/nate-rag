"""Parse CSV / XLSX uploads into QA rows using user-selected columns."""
import csv
import io
import logging
from typing import Any

from backend.app.services.ingestion.qa_models import QaPairInput

logger = logging.getLogger(__name__)


def _assert_columns(columns: list[str], question_column: str, answer_column: str) -> None:
    if question_column not in columns:
        raise ValueError(
            f"Question column {question_column!r} not in file. "
            f"Available: {', '.join(columns)}"
        )
    if answer_column not in columns:
        raise ValueError(
            f"Answer column {answer_column!r} not in file. "
            f"Available: {', '.join(columns)}"
        )


def _split_tags(raw: Any) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    for sep in (";", "|"):
        if sep in text:
            return [t.strip() for t in text.split(sep) if t.strip()]
    if "," in text:
        return [t.strip() for t in text.split(",") if t.strip()]
    return [text]


def _row_get(row: dict[str, Any], column: str) -> str:
    if column not in row:
        raise ValueError(f"Column not found: {column!r}")
    val = row[column]
    if val is None:
        return ""
    return str(val).strip()


def rows_from_dicts(
    records: list[dict[str, Any]],
    *,
    question_column: str,
    answer_column: str,
    tags_column: str | None = None,
    document_name_column: str | None = None,
    default_document_name: str = "Unknown",
    source: str = "spreadsheet",
) -> tuple[list[QaPairInput], int, int]:
    """Returns (valid_pairs, total_rows, skipped_empty)."""
    q_col = question_column.strip()
    a_col = answer_column.strip()
    t_col = tags_column.strip() if tags_column else None
    d_col = document_name_column.strip() if document_name_column else None

    pairs: list[QaPairInput] = []
    skipped = 0

    for row in records:
        question = _row_get(row, q_col)
        answer = _row_get(row, a_col)
        if not question or not answer:
            skipped += 1
            continue

        doc_name = default_document_name
        if d_col and d_col in row and str(row[d_col]).strip():
            doc_name = str(row[d_col]).strip()

        tags = _split_tags(row.get(t_col)) if t_col and t_col in row else []

        pairs.append(
            QaPairInput(
                question=question,
                answer=answer,
                tags=tags,
                document_name=doc_name,
                source=source,  # type: ignore[arg-type]
            )
        )

    return pairs, len(records), skipped


def parse_csv_bytes(
    raw: bytes,
    *,
    question_column: str,
    answer_column: str,
    tags_column: str | None = None,
    document_name_column: str | None = None,
    default_document_name: str = "Unknown",
) -> tuple[list[QaPairInput], int, int, list[str]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV has no header row")
    columns = list(reader.fieldnames)
    _assert_columns(columns, question_column, answer_column)
    records = list(reader)
    pairs, total, skipped = rows_from_dicts(
        records,
        question_column=question_column,
        answer_column=answer_column,
        tags_column=tags_column,
        document_name_column=document_name_column,
        default_document_name=default_document_name,
    )
    return pairs, total, skipped, columns


def parse_xlsx_bytes(
    raw: bytes,
    *,
    question_column: str,
    answer_column: str,
    tags_column: str | None = None,
    document_name_column: str | None = None,
    default_document_name: str = "Unknown",
) -> tuple[list[QaPairInput], int, int, list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError(
            "XLSX support requires openpyxl. Install with: pip install openpyxl"
        ) from e

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as e:
        raise ValueError("Spreadsheet is empty") from e

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    if not any(headers):
        raise ValueError("Spreadsheet has no header row")

    _assert_columns(headers, question_column, answer_column)

    records: list[dict[str, Any]] = []
    for row_values in rows_iter:
        if row_values is None or all(v is None or str(v).strip() == "" for v in row_values):
            continue
        record = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            if i < len(row_values):
                record[header] = row_values[i]
        if record:
            records.append(record)

    wb.close()

    pairs, total, skipped = rows_from_dicts(
        records,
        question_column=question_column,
        answer_column=answer_column,
        tags_column=tags_column,
        document_name_column=document_name_column,
        default_document_name=default_document_name,
    )
    return pairs, total, skipped, headers


def parse_spreadsheet_bytes(
    raw: bytes,
    filename: str,
    *,
    question_column: str,
    answer_column: str,
    tags_column: str | None = None,
    document_name_column: str | None = None,
    default_document_name: str = "Unknown",
) -> tuple[list[QaPairInput], int, int, list[str]]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return parse_csv_bytes(
            raw,
            question_column=question_column,
            answer_column=answer_column,
            tags_column=tags_column,
            document_name_column=document_name_column,
            default_document_name=default_document_name,
        )
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return parse_xlsx_bytes(
            raw,
            question_column=question_column,
            answer_column=answer_column,
            tags_column=tags_column,
            document_name_column=document_name_column,
            default_document_name=default_document_name,
        )
    raise ValueError("File must be .csv or .xlsx")


def spreadsheet_columns(raw: bytes, filename: str) -> list[str]:
    """Return header column names without parsing row data."""
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV has no header row")
        return list(reader.fieldnames)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ValueError(
                "XLSX support requires openpyxl. Install with: pip install openpyxl"
            ) from e
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as e:
            raise ValueError("Spreadsheet is empty") from e
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        wb.close()
        if not any(headers):
            raise ValueError("Spreadsheet has no header row")
        return [h for h in headers if h]
    raise ValueError("File must be .csv or .xlsx")
